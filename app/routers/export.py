"""
Router للتصدير - Export Router (PDF & Excel)
"""
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
from datetime import date, datetime, timedelta
from decimal import Decimal
from typing import Optional
import io

from ..database import get_db
from ..utils.dependencies import get_current_user
from ..models import User, Booking, Unit, Customer, Owner, Project, Transaction


router = APIRouter(prefix="/api/export", tags=["Export"])


# ============ PDF Export ============

def generate_pdf_report(
    title: str,
    period_label: str,
    summary_data: dict,
    table_data: list,
    table_headers: list
) -> bytes:
    """
    توليد تقرير PDF باستخدام ReportLab مع دعم اللغة العربية
    """
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm, mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    import os
    
    # Try to use arabic-reshaper and bidi for proper Arabic text
    try:
        import arabic_reshaper
        from bidi.algorithm import get_display
        HAS_ARABIC_SUPPORT = True
    except ImportError:
        HAS_ARABIC_SUPPORT = False
    
    def reshape_arabic(text):
        """Reshape Arabic text for proper display"""
        if not HAS_ARABIC_SUPPORT or not text:
            return text
        try:
            reshaped = arabic_reshaper.reshape(str(text))
            return get_display(reshaped)
        except:
            return text
    
    # Register Arabic font - try multiple fallbacks
    font_registered = False
    font_name = 'Helvetica'  # fallback
    
    # List of possible Arabic font paths
    arabic_fonts = [
        # Windows fonts
        "C:/Windows/Fonts/arial.ttf",
        "C:/Windows/Fonts/tahoma.ttf",
        "C:/Windows/Fonts/calibri.ttf",
        # Linux fonts
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        "/usr/share/fonts/truetype/freefont/FreeSans.ttf",
    ]
    
    for font_path in arabic_fonts:
        if os.path.exists(font_path):
            try:
                pdfmetrics.registerFont(TTFont('ArabicFont', font_path))
                font_name = 'ArabicFont'
                font_registered = True
                break
            except:
                continue
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=2*cm,
        leftMargin=2*cm,
        topMargin=2*cm,
        bottomMargin=2*cm
    )
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Title Style
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontName=font_name,
        fontSize=24,
        spaceAfter=30,
        alignment=1,  # Center
        textColor=colors.HexColor('#1e293b')
    )
    
    # Subtitle Style
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Normal'],
        fontName=font_name,
        fontSize=12,
        spaceAfter=20,
        alignment=1,
        textColor=colors.HexColor('#64748b')
    )
    
    # Header
    elements.append(Paragraph(reshape_arabic(f"🏠 {title}"), title_style))
    elements.append(Paragraph(reshape_arabic(period_label), subtitle_style))
    elements.append(Paragraph(reshape_arabic(f"تاريخ التصدير: {datetime.now().strftime('%Y-%m-%d %H:%M')}"), subtitle_style))
    elements.append(Spacer(1, 20))
    
    # Summary Section
    if summary_data:
        summary_title = ParagraphStyle(
            'SummaryTitle',
            parent=styles['Heading2'],
            fontName=font_name,
            fontSize=16,
            spaceAfter=10,
            textColor=colors.HexColor('#1e293b')
        )
        elements.append(Paragraph(reshape_arabic("📊 ملخص التقرير"), summary_title))
        
        summary_items = [[reshape_arabic(k), reshape_arabic(str(v))] for k, v in summary_data.items()]
        summary_table = Table(summary_items, colWidths=[8*cm, 6*cm])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#f8fafc')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#1e293b')),
            ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, -1), font_name),
            ('FONTSIZE', (0, 0), (-1, -1), 11),
            ('PADDING', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
            ('ROWBACKGROUNDS', (0, 0), (-1, -1), [colors.HexColor('#f8fafc'), colors.white]),
        ]))
        elements.append(summary_table)
        elements.append(Spacer(1, 30))
    
    # Data Table
    if table_data and table_headers:
        table_title = ParagraphStyle(
            'TableTitle',
            parent=styles['Heading2'],
            fontName=font_name,
            fontSize=16,
            spaceAfter=10,
            textColor=colors.HexColor('#1e293b')
        )
        elements.append(Paragraph(reshape_arabic("📋 البيانات التفصيلية"), table_title))
        
        # Reshape headers and data for Arabic
        reshaped_headers = [reshape_arabic(h) for h in table_headers]
        reshaped_data = [[reshape_arabic(str(cell)) for cell in row] for row in table_data]
        
        # Prepare table with headers
        full_data = [reshaped_headers] + reshaped_data
        
        # Calculate column widths
        col_count = len(table_headers)
        col_width = (A4[0] - 4*cm) / col_count
        
        data_table = Table(full_data, colWidths=[col_width] * col_count)
        data_table.setStyle(TableStyle([
            # Header
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('FONTNAME', (0, 0), (-1, -1), font_name),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            # Body
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.HexColor('#1e293b')),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('ALIGN', (0, 1), (-1, -1), 'RIGHT'),
            # General
            ('PADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')]),
        ]))
        elements.append(data_table)
    
    # Build PDF
    doc.build(elements)
    buffer.seek(0)
    return buffer.getvalue()


@router.get("/pdf/financial")
@router.get("/pdf/financial/")
async def export_financial_pdf(
    period: str = Query("monthly", description="daily, weekly, monthly"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """تصدير تقرير مالي PDF"""
    today = date.today()
    
    if period == "daily":
        start_date = today
        period_label = f"تقرير يومي - {today.strftime('%Y-%m-%d')}"
    elif period == "weekly":
        start_date = today - timedelta(days=today.weekday())
        period_label = f"تقرير أسبوعي - من {start_date.strftime('%Y-%m-%d')} إلى {today.strftime('%Y-%m-%d')}"
    else:  # monthly
        start_date = date(today.year, today.month, 1)
        period_label = f"تقرير شهري - {today.strftime('%Y-%m')}"
    
    # الحجوزات
    bookings = db.query(Booking).filter(
        Booking.check_in_date >= start_date,
        Booking.status.in_(["مؤكد", "دخول", "مكتمل"]),
        Booking.is_deleted == False
    ).all()
    
    total_bookings = len(bookings)
    total_revenue = sum(float(b.total_price or 0) for b in bookings)
    
    # الإلغاءات
    cancellations = db.query(Booking).filter(
        func.date(Booking.created_at) >= start_date,
        Booking.status == "ملغي"
    ).count()
    
    # المعاملات المالية
    transactions = db.query(Transaction).filter(
        Transaction.date >= start_date
    ).all()
    
    income = sum(float(t.amount) for t in transactions if t.type == "دخل")
    expenses = sum(float(t.amount) for t in transactions if t.type == "صرف")
    
    # Summary data
    summary_data = {
        "إجمالي الحجوزات": total_bookings,
        "إجمالي الإيرادات": f"{total_revenue:,.2f} ر.س",
        "الإلغاءات": cancellations,
        "الدخل من المعاملات": f"{income:,.2f} ر.س",
        "المصروفات": f"{expenses:,.2f} ر.س",
        "صافي الربح": f"{income - expenses:,.2f} ر.س",
    }
    
    # Table data
    table_headers = ["اسم الضيف", "تاريخ الوصول", "تاريخ المغادرة", "المبلغ", "الحالة"]
    table_data = [
        [
            b.guest_name,
            b.check_in_date.strftime('%Y-%m-%d') if b.check_in_date else '-',
            b.check_out_date.strftime('%Y-%m-%d') if b.check_out_date else '-',
            f"{float(b.total_price or 0):,.2f}",
            b.status
        ]
        for b in bookings[:50]  # Limit to 50
    ]
    
    # Generate PDF
    pdf_bytes = generate_pdf_report(
        title="تقرير مالي - نظام منام",
        period_label=period_label,
        summary_data=summary_data,
        table_data=table_data,
        table_headers=table_headers
    )
    
    filename = f"financial_report_{period}_{today.strftime('%Y%m%d')}.pdf"
    
    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ============ Excel Export ============

def generate_excel_file(
    sheet_name: str,
    headers: list,
    data: list,
    title: str = None
) -> bytes:
    """
    توليد ملف Excel باستخدام openpyxl
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.sheet_view.rightToLeft = True  # RTL for Arabic
    
    row_num = 1
    
    # Title row (if provided)
    if title:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        title_cell = ws.cell(row=1, column=1, value=title)
        title_cell.font = Font(size=16, bold=True, color="1e293b")
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        row_num = 3
    
    # Header row
    header_fill = PatternFill(start_color="3b82f6", end_color="3b82f6", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=row_num, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    row_num += 1
    
    # Data rows
    data_alignment = Alignment(horizontal='right', vertical='center')
    alt_fill = PatternFill(start_color="f8fafc", end_color="f8fafc", fill_type="solid")
    
    for row_idx, row_data in enumerate(data):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.alignment = data_alignment
            if row_idx % 2 == 1:
                cell.fill = alt_fill
        row_num += 1
    
    # Auto-adjust column widths
    for col_num in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_num)
        max_length = max(
            len(str(headers[col_num - 1])),
            max((len(str(row[col_num - 1])) for row in data), default=0)
        )
        ws.column_dimensions[column_letter].width = min(max_length + 4, 50)
    
    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


@router.get("/excel/bookings")
@router.get("/excel/bookings/")
async def export_bookings_excel(
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """تصدير الحجوزات Excel"""
    query = db.query(Booking).filter(Booking.is_deleted == False)
    
    if start_date:
        query = query.filter(Booking.check_in_date >= start_date)
    if end_date:
        query = query.filter(Booking.check_in_date <= end_date)
    
    bookings = query.order_by(Booking.check_in_date.desc()).all()
    
    headers = ["اسم الضيف", "رقم الجوال", "تاريخ الوصول", "تاريخ المغادرة", "المبلغ", "الحالة", "المصدر"]
    data = [
        [
            b.guest_name,
            b.guest_phone or '-',
            b.check_in_date.strftime('%Y-%m-%d') if b.check_in_date else '-',
            b.check_out_date.strftime('%Y-%m-%d') if b.check_out_date else '-',
            float(b.total_price or 0),
            b.status,
            b.channel_source or 'مباشر'
        ]
        for b in bookings
    ]
    
    excel_bytes = generate_excel_file(
        sheet_name="الحجوزات",
        headers=headers,
        data=data,
        title=f"قائمة الحجوزات - {datetime.now().strftime('%Y-%m-%d')}"
    )
    
    filename = f"bookings_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        io.BytesIO(excel_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/excel/units")
@router.get("/excel/units/")
async def export_units_excel(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """تصدير الوحدات Excel"""
    units = db.query(Unit).filter(Unit.is_deleted == False).all()
    
    headers = ["اسم الوحدة", "النوع", "الغرف", "الحالة", "سعر أيام الأسبوع", "سعر نهاية الأسبوع"]
    data = [
        [
            u.unit_name,
            u.unit_type,
            u.rooms,
            u.status,
            float(u.price_days_of_week or 0),
            float(u.price_in_weekends or 0)
        ]
        for u in units
    ]
    
    excel_bytes = generate_excel_file(
        sheet_name="الوحدات",
        headers=headers,
        data=data,
        title=f"قائمة الوحدات - {datetime.now().strftime('%Y-%m-%d')}"
    )
    
    filename = f"units_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        io.BytesIO(excel_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/excel/customers")
@router.get("/excel/customers/")
async def export_customers_excel(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """تصدير العملاء Excel"""
    customers = db.query(Customer).filter(Customer.is_deleted == False).all()
    
    headers = ["الاسم", "رقم الجوال", "البريد", "عدد الحجوزات", "إجمالي الإيراد", "الحالة"]
    data = [
        [
            c.name,
            c.phone,
            c.email or '-',
            c.booking_count,
            float(c.total_revenue or 0),
            "محظور" if c.is_banned else "نشط"
        ]
        for c in customers
    ]
    
    excel_bytes = generate_excel_file(
        sheet_name="العملاء",
        headers=headers,
        data=data,
        title=f"قائمة العملاء - {datetime.now().strftime('%Y-%m-%d')}"
    )
    
    filename = f"customers_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        io.BytesIO(excel_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/excel/transactions")
@router.get("/excel/transactions/")
async def export_transactions_excel(
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """تصدير المعاملات المالية Excel"""
    query = db.query(Transaction)
    
    if start_date:
        query = query.filter(Transaction.date >= start_date)
    if end_date:
        query = query.filter(Transaction.date <= end_date)
    
    transactions = query.order_by(Transaction.date.desc()).all()
    
    headers = ["التاريخ", "النوع", "المبلغ", "الوصف", "الفئة"]
    data = [
        [
            t.date.strftime('%Y-%m-%d') if t.date else '-',
            t.type,
            float(t.amount or 0),
            t.description or '-',
            t.category or '-'
        ]
        for t in transactions
    ]
    
    excel_bytes = generate_excel_file(
        sheet_name="المعاملات",
        headers=headers,
        data=data,
        title=f"المعاملات المالية - {datetime.now().strftime('%Y-%m-%d')}"
    )
    
    filename = f"transactions_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        io.BytesIO(excel_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/excel/owners")
@router.get("/excel/owners/")
async def export_owners_excel(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """تصدير الملاك Excel"""
    owners = db.query(Owner).filter(Owner.is_deleted == False).all()
    
    headers = ["اسم المالك", "رقم الجوال", "البريد", "عدد المشاريع", "ملاحظات"]
    data = [
        [
            o.owner_name,
            o.owner_mobile_phone,
            o.paypal_email or '-',
            len(o.projects) if o.projects else 0,
            o.note or '-'
        ]
        for o in owners
    ]
    
    excel_bytes = generate_excel_file(
        sheet_name="الملاك",
        headers=headers,
        data=data,
        title=f"قائمة الملاك - {datetime.now().strftime('%Y-%m-%d')}"
    )
    
    filename = f"owners_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        io.BytesIO(excel_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
